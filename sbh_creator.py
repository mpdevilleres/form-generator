#!/usr/bin/env python
import copy

import re

import os
import datetime as dt

from django.db.models import Count, Q
from django.core.wsgi import get_wsgi_application

BASE_DIR = os.path.dirname(os.path.realpath(__file__))
FORM_DIR = os.path.join(BASE_DIR, "forms")

os.environ.setdefault("DJANGO_SETTINGS_MODULE", "project.settings")

application = get_wsgi_application()

from main.models import PurchaseOrder, PurchaseOrderLineDetail, Resource, UnitPrice, Invoice
from dump import to_date_format
import pandas as pd
from openpyxl import *
from openpyxl.cell import Cell
from openpyxl.utils import get_column_letter
from openpyxl.worksheet import Worksheet


def insert_rows(self, row_idx, cnt, above=False, copy_style=True, fill_formulae=True):
    """Inserts new (empty) rows into worksheet at specified row index.

    :param self: Class object
    :param row_idx: Row index specifying where to insert new rows.
    :param cnt: Number of rows to insert.
    :param above: Set True to insert rows above specified row index.
    :param copy_style: Set True if new rows should copy style of immediately above row.
    :param fill_formulae: Set True if new rows should take on formula from immediately above row, filled with references new to rows.

    Usage:

    * insert_rows(2, 10, above=True, copy_style=False)

    """
    CELL_RE = re.compile("(?P<col>\$?[A-Z]+)(?P<row>\$?\d+)")

    row_idx = row_idx - 1 if above else row_idx

    def replace(m):
        row = m.group('row')
        prefix = "$" if row.find("$") != -1 else ""
        row = int(row.replace("$", ""))
        row += cnt if row > row_idx else 0
        return m.group('col') + prefix + str(row)

    # First, we shift all cells down cnt rows...
    old_cells = set()
    old_fas = set()
    new_cells = dict()
    new_fas = dict()
    for c in self._cells.values():

        old_coor = c.coordinate

        # Shift all references to anything below row_idx
        if c.data_type == Cell.TYPE_FORMULA:
            c.value = CELL_RE.sub(
                replace,
                c.value
            )
            # Here, we need to properly update the formula references to reflect new row indices
            if old_coor in self.formula_attributes and 'ref' in self.formula_attributes[old_coor]:
                self.formula_attributes[old_coor]['ref'] = CELL_RE.sub(
                    replace,
                    self.formula_attributes[old_coor]['ref']
                )

        # Do the magic to set up our actual shift
        if c.row > row_idx:
            old_coor = c.coordinate
            old_cells.add((c.row, c.col_idx))
            c.row += cnt
            new_cells[(c.row, c.col_idx)] = c
            if old_coor in self.formula_attributes:
                old_fas.add(old_coor)
                fa = self.formula_attributes[old_coor].copy()
                new_fas[c.coordinate] = fa

    for coor in old_cells:
        del self._cells[coor]
    self._cells.update(new_cells)

    for fa in old_fas:
        del self.formula_attributes[fa]
    self.formula_attributes.update(new_fas)

    # Next, we need to shift all the Row Dimensions below our new rows down by cnt...
    for row in range(len(self.row_dimensions) - 1 + cnt, row_idx + cnt, -1):
        new_rd = copy.copy(self.row_dimensions[row - cnt])
        new_rd.index = row
        self.row_dimensions[row] = new_rd
        del self.row_dimensions[row - cnt]

    # Now, create our new rows, with all the pretty cells
    row_idx += 1
    for row in range(row_idx, row_idx + cnt):
        # Create a Row Dimension for our new row
        new_rd = copy.copy(self.row_dimensions[row - 1])
        new_rd.index = row
        self.row_dimensions[row] = new_rd
        for col in range(1, self.max_column):
            col = get_column_letter(col)
            cell = self.cell('%s%d' % (col, row))
            cell.value = None
            source = self.cell('%s%d' % (col, row - 1))
            if copy_style:
                cell.number_format = source.number_format
                cell.font = source.font.copy()
                cell.alignment = source.alignment.copy()
                cell.border = source.border.copy()
                cell.fill = source.fill.copy()
            if fill_formulae and source.data_type == Cell.TYPE_FORMULA:
                s_coor = source.coordinate
                if s_coor in self.formula_attributes and 'ref' not in self.formula_attributes[s_coor]:
                    fa = self.formula_attributes[s_coor].copy()
                    self.formula_attributes[cell.coordinate] = fa
                # print("Copying formula from cell %s%d to %s%d"%(col,row-1,col,row))
                cell.value = re.sub(
                    "(\$?[A-Z]{1,3}\$?)%d" % (row - 1),
                    lambda m: m.group(1) + str(row),
                    source.value
                )
                cell.data_type = Cell.TYPE_FORMULA

    # Check for Merged Cell Ranges that need to be expanded to contain new cells
    for cr_idx, cr in enumerate(self.merged_cell_ranges):
        self.merged_cell_ranges[cr_idx] = CELL_RE.sub(
            replace,
            cr
        )


Worksheet.insert_rows = insert_rows


# MAIN COMPONENTS STARTS HERE
# ----------------------------------------------------------------------------------------------------
def get_required_hour(period_start=None, period_end=None, ramadan_start=None, ramadan_end=None):
    ramadan_days = 0
    if ramadan_start is None and ramadan_end is None:
        delta = period_end - period_start
        normal_days = delta.days + 1
    else:
        start = period_start
        if period_start < ramadan_start:
            start = ramadan_start
        end = period_end
        if ramadan_end < period_end:
            end = ramadan_end

        total = period_end - period_start
        ramadan = end - start
        total_days = total.days
        ramadan_days = ramadan.days + 1
        normal_days = total_days - ramadan_days

    # +1 is a corrections factor to count all days within a period
    return round(normal_days * 48 / 7) + round(ramadan_days * 36 / 7)


def get_individual_required_hour(date_of_join=None, period_start=None, period_end=None,
                                 ramadan_start=None, ramadan_end=None):
    # date of date is latest
    if date_of_join > period_start:
        return get_required_hour(date_of_join, period_end, ramadan_start, ramadan_end)
    return get_required_hour(period_start, period_end, ramadan_start, ramadan_end)


class SBH(object):
    period_start = None
    period_end = None
    ramadan_start = None
    ramadan_end = None
    period_string = None
    required_hour = None
    po_num = None
    context = None
    output_file = None

    def __init__(self):
        self.base_dir = os.path.dirname(os.path.realpath(__file__))
        self.form = os.path.join(self.base_dir, 'forms', 'SBH-FORM.xlsx')

    def set_variables(self):
        self.period_start = to_date_format(self.period_start)
        self.period_end = to_date_format(self.period_end)
        self.ramadan_start = None if self.ramadan_start is None else to_date_format(self.ramadan_start)
        self.ramadan_end = None if self.ramadan_end is None else to_date_format(self.ramadan_end)
        self.period_string = '{0:%d-%b-%Y} to {1:%d-%b-%Y}'.format(self.period_start,
                                                                   self.period_end)
        self.required_hour = get_required_hour(self.period_start, self.period_end,
                                               self.ramadan_start, self.ramadan_end)

    def get_context(self):
        # PERIOD MUST BE IN dd/mm/yyyy format()
        self.required_hour = get_required_hour(self.period_start, self.period_end, self.ramadan_start,
                                               self.ramadan_end)
        # obj is a single PurchaseOrder model
        q = Q(po_num=self.po_num)
        qs_po = PurchaseOrder.objects.filter(q). \
            annotate(line=Count('purchaseorderline__pk')). \
            first()

        # queryset for all resources filtered by po_num
        # use for rs_resource
        qs_resource = Resource.objects.filter(q)
        q = Q()

        for obj in qs_resource.all():
            q |= Q(resource__pk=obj.pk)

        qs_invoice = Invoice.objects.order_by('-invoice_date').filter(invoice_date=dt.datetime(self.period_start.year,
                                                                                               self.period_start.month,
                                                                                               1)
                                                                      )

        df_resource = pd.DataFrame(list(qs_resource.values('id', 'po_os_ref', 'agency_ref_num', 'res_full_name',
                                                           'po_position', 'po_level', 'date_of_join',
                                                           'po_line_detail__director_name',
                                                           'po_line_detail__rate_diff_percent'
                                                           )
                                        )
                                   )
        df_resource['period_start'] = self.period_start.date()
        df_resource['period_end'] = self.period_end.date()
        df_resource['required_hour'] = self.required_hour
        # df_resource['required_hour'] = df_resource.apply(lambda row: get_individual_required_hour(row['date_of_join'],
        #                                                                                           row['period_start'],
        #                                                                                           row['period_end']),
        #                                                  axis=1)

        df_invoice = pd.DataFrame(list(qs_invoice.values('id', 'resource_id', 'invoice_hour', 'invoice_claim',
                                                         'remarks'
                                                         )
                                       )
                                  )
        if df_invoice.empty:
            df_invoice = pd.DataFrame(columns=['resource_id', 'id', 'invoice_claim', 'invoice_hour', 'remarks'])
        else:
            df_invoice = df_invoice.groupby('resource_id').first().reset_index()

        rs_resource = pd.merge(left=df_resource, right=df_invoice, how='left', left_on='id', right_on='resource_id')

        rs_resource = rs_resource.fillna(0.0).to_dict('records')

        # queryset for all PO Line filtered by po_num
        q = Q()
        for obj in qs_po.purchaseorderline_set.all():
            q |= Q(po_line__pk=obj.pk)
        qs_line_details = PurchaseOrderLineDetail.objects.filter(q)

        # to get rs_summary

        df_line_details = pd.DataFrame(list(qs_line_details.values('po_position', 'po_os_ref', 'po_level',
                                                                   'rate_diff_percent')))
        df_line_details = df_line_details.groupby(['po_os_ref', 'po_position', 'po_level', 'rate_diff_percent']).\
            size().reset_index()
        df_line_details.columns = ['po_os_ref', 'po_position', 'po_level', 'rate_diff_percent', 'count']
        df_line_details = df_line_details.pivot_table(index=['po_position', 'po_os_ref', 'rate_diff_percent'],
                                                      columns='po_level',
                                                      values='count')

        rs_summary = df_line_details.reset_index().fillna(0.0).to_dict('records')

        # get unit price
        q = Q(contractor=qs_po.contractor)
        qs_unit_price = UnitPrice.objects.filter(q)
        df_unit_price = pd.DataFrame(list(qs_unit_price.values()))
        df_unit_price = df_unit_price.pivot(index='po_position', columns='po_level', values='amount').reset_index()
        rs_unit_price = df_unit_price.to_dict('records')

        context = {
            'contractor': qs_po.contractor,
            'po_num': qs_po.po_num,
            'po_line_count': qs_po.line,
            'rs_resource': rs_resource,  # recordset for individual resource (employee information)
            'rs_summary': rs_summary,  # recordset for summary of total numbers
            'rs_unit_price': rs_unit_price,  # recordset for summary of unit price
            'period_string': self.period_string,
            'total_required_hour': self.required_hour
        }

        return context

    def get_output_file(self, context):
        return "{:0>2} {} {} {}.xlsx".format(self.period_start.month,
                                      context['contractor'],
                                      context['po_num'],
                                      context['period_string'])

    def single_set_sbh(self, context=None):

        wb = load_workbook(self.form)

        # WORK SHEET SBH-FORM 1
        # ---------------------------------------------------------------------------------------
        row = 15
        column = 2
        sr = 1
        ws = wb.get_sheet_by_name('SBH-FORM 1')

        ws.cell('F4').value = context['contractor']
        ws.cell('K4').value = context['po_num']
        ws.cell('K6').value = context['po_line_count']
        ws.cell('F6').value = context['period_string']

        rs_resource = context['rs_resource']

        ws.insert_rows(row, len(rs_resource) - 1)

        for record in rs_resource:
            ws.cell(row=row, column=column).value = sr
            ws.cell(row=row, column=column + 1).value = record['po_os_ref']
            ws.cell(row=row, column=column + 2).value = record['agency_ref_num']
            ws.cell(row=row, column=column + 3).value = record['po_line_detail__director_name']
            ws.cell(row=row, column=column + 4).value = record['res_full_name']
            ws.cell(row=row, column=column + 5).value = record['po_position']
            ws.cell(row=row, column=column + 6).value = record['po_level']
            ws.cell(row=row, column=column + 7).value = '{0:%d-%b-%Y}'.format(record['date_of_join'])
            ws.cell(row=row, column=column + 8).value = record['po_line_detail__rate_diff_percent']
            ws.cell(row=row, column=column + 9).value = record['required_hour']
            ws.cell(row=row, column=column + 10).value = record['invoice_claim']
            ws.cell(row=row, column=column + 12).value = record['remarks']
            sr += 1
            row += 1
        # ---------------------------------------------------------------------------------------
        # END SBH FORM 1


        # WORK SHEET SBH-FORM 2
        # ---------------------------------------------------------------------------------------
        row = 16
        column = 2  # B

        ws = wb.get_sheet_by_name('SBH-FORM 2')

        ws.insert_rows(row, len(context['rs_summary']) - 1)
        ws.cell('P13').value = context['total_required_hour']

        for record in context['rs_summary']:
            ws.cell(row=row, column=column).value = record['po_os_ref']
            ws.cell(row=row, column=column + 1).value = record.get('po_position')
            ws.cell(row=row, column=column + 2).value = record.get('Level 1', 0)
            ws.cell(row=row, column=column + 3).value = record.get('Level 2', 0)
            ws.cell(row=row, column=column + 4).value = record.get('Level 3', 0)
            ws.cell(row=row, column=column + 12).value = record.get('rate_diff_percent', None)

            row += 1
        # ---------------------------------------------------------------------------------------
        # END SBH FORM 2

        # WORK SHEET UNIT PRICE
        # ---------------------------------------------------------------------------------------
        row = 1
        column = 1  # B

        ws = wb.get_sheet_by_name('UNIT PRICE')

        for record in context['rs_unit_price']:
            ws.cell(row=row, column=column).value = record['po_position']
            ws.cell(row=row, column=column + 1).value = record.get('Level 1', 0)
            ws.cell(row=row, column=column + 2).value = record.get('Level 2', 0)
            ws.cell(row=row, column=column + 3).value = record.get('Level 3', 0)
            ws.cell(row=row, column=column + 4).value = record.get('Level 4', 0)

            row += 1
        # ---------------------------------------------------------------------------------------
        # END UNIT PRICE

        output_file = self.get_output_file(context)
        wb.save(os.path.join(BASE_DIR, 'output', output_file))

    def make_sbh_per_po(self, po_num=None, period_start=None, period_end=None,
                        ramadan_start=None, ramadan_end=None):
        self.po_num = po_num
        self.period_start = period_start
        self.period_end = period_end
        self.ramadan_start = ramadan_start
        self.ramadan_end = ramadan_end
        self.set_variables()
        try:
            context = self.get_context()
            self.single_set_sbh(context)
        except Exception as e:
            print('Failed: ', po_num, ' due to missing {}'.format(str(e)))

    def make_sbh_per_contractor(self, contractor=None, period_start=None, period_end=None,
                                  ramadan_start=None, ramadan_end=None):
        if contractor.lower() == 'all':
            qs = PurchaseOrder.objects.all()
        else:
            qs = PurchaseOrder.objects.filter(contractor=contractor).all()

        for i in qs:
                self.make_sbh_per_po(i.po_num, period_start, period_end, ramadan_start, ramadan_end)

if __name__ == '__main__':
    pass
    # sbh = SBH()
    # sbh.make_forms_per_contractor('REACH', "20/06/2016", "20/07/2016", '20/06/2016', '05/07/2016')
