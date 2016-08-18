#!/usr/bin/env python

# IMPORT INITIALIZING DJANGO ORM
# IMPORT OPENPYXL WITH INSERT ROW
# excel password tbpc19
# ----------------------------------------------------------------------------------------------------
from django.db.models import Q, Count
from openpyxl.styles import Protection

from utils import *
from utils import load_workbook
from main.models import PurchaseOrder, PurchaseOrderLineDetail, Resource, UnitPrice, Invoice
import pandas as pd


# MAIN COMPONENTS STARTS HERE
# ----------------------------------------------------------------------------------------------------
def get_required_hour(period_start=None, period_end=None, ramadan_start=None, ramadan_end=None):
    total_days = (period_end - period_start).days + 1
    ramadan_days = 0

    if ramadan_start and ramadan_end:
        start = period_start
        if period_start < ramadan_start:
            start = ramadan_start
        end = period_end
        if ramadan_end < period_end:
            end = ramadan_end

        ramadan = end - start
        ramadan_days = ramadan.days

    normal_days = total_days - ramadan_days

    # +1 is a corrections factor to count all days within a period
    return round(normal_days * 48 / 7) + round(ramadan_days * 36 / 7)


def get_individual_required_hour(date_of_join=None, period_start=None, period_end=None,
                                 ramadan_start=None, ramadan_end=None):
    # date of date is latest
    if date_of_join > period_start:
        return get_required_hour(date_of_join, period_end, ramadan_start, ramadan_end)
    return get_required_hour(period_start, period_end, ramadan_start, ramadan_end)


def get_output_file(*args):
    return ' '.join(args) + '.xlsx'


class SBH(object):
    period_start = None
    period_end = None
    ramadan_start = None
    ramadan_end = None
    period_string = None
    required_hour = None
    po_num = None
    context = None
    output_file = None
    qs_po = None
    qs_resource = None
    qs_invoice = None
    qs_line_details = None
    qs_unit_price = None
    division = None

    def __init__(self):
        self.base_dir = os.path.dirname(os.path.realpath(__file__))
        self.form = os.path.join(self.base_dir, 'forms', 'SBH-FORM.xlsx')

    def set_variables(self):
        self.period_start = to_date_format(self.period_start)
        self.period_end = to_date_format(self.period_end)
        self.ramadan_start = None if self.ramadan_start is None else to_date_format(self.ramadan_start)
        self.ramadan_end = None if self.ramadan_end is None else to_date_format(self.ramadan_end)
        self.period_string = '{0:%d-%b-%Y} to {1:%d-%b-%Y}'.format(self.period_start,
                                                                   self.period_end)
        # PERIOD MUST BE IN dd/mm/yyyy format()
        self.required_hour = get_required_hour(self.period_start, self.period_end,
                                               self.ramadan_start, self.ramadan_end)

    def set_initial_queryset(self):
        q = Q(po_num=self.po_num)
        self.qs_po = PurchaseOrder.objects.filter(q). \
            annotate(line=Count('purchaseorderline__pk')). \
            first()

        # queryset for all resources filtered by po_num
        # RPT01_ Monthly Accruals - Mobillized

        self.qs_resource = Resource.objects.filter(q)
        if self.division:
            self.qs_resource = self.qs_resource.filter(division=self.division)

        q = Q()
        for obj in self.qs_resource.all():
            q |= Q(resource__pk=obj.pk)

        self.qs_invoice = Invoice.objects.order_by('-invoice_date').filter(
            invoice_date=dt.datetime(self.period_start.year,
                                     self.period_start.month,
                                     1)
        )

        # queryset for all PO Line filtered by po_num
        q = Q()
        for obj in self.qs_po.purchaseorderline_set.all():
            q |= Q(po_line__pk=obj.pk)
        self.qs_line_details = PurchaseOrderLineDetail.objects.filter(q)
        if self.division:
            self.qs_line_details = self.qs_line_details.filter(division=self.division)

        # get unit price
        q = Q(contractor=self.qs_po.contractor)
        self.qs_unit_price = UnitPrice.objects.filter(q)

    def get_context(self):
        # obj is a single PurchaseOrder model

        df_resource = pd.DataFrame(list(self.qs_resource.values('id', 'po_os_ref', 'agency_ref_num', 'res_full_name',
                                                                'po_position', 'po_level', 'date_of_join',
                                                                'po_line_detail__director_name',
                                                                'po_line_detail__rate_diff_percent'
                                                                )
                                        )
                                   )
        df_resource['period_start'] = self.period_start.date()
        df_resource['period_end'] = self.period_end.date()
        df_resource['required_hour'] = self.required_hour

        df_invoice = pd.DataFrame(list(self.qs_invoice.values('id', 'resource_id', 'invoice_hour', 'invoice_claim',
                                                              'remarks'
                                                              )
                                       )
                                  )
        if df_invoice.empty:
            df_invoice = pd.DataFrame(columns=['resource_id', 'id', 'invoice_claim', 'invoice_hour', 'remarks'])
        else:
            df_invoice = df_invoice.groupby('resource_id').first().reset_index()

        rs_resource = pd.merge(left=df_resource, right=df_invoice, how='left', left_on='id', right_on='resource_id')
        rs_resource.sort_values(
            by=['po_line_detail__director_name', 'po_position', 'po_line_detail__rate_diff_percent'],
            ascending=[True, True, True],
            inplace=True)
        rs_resource = rs_resource.fillna(0.0).to_dict('records')

        # to get rs_summary

        df_line_details = pd.DataFrame(list(self.qs_line_details.values('po_position', 'po_os_ref', 'po_level',
                                                                        'rate_diff_percent')))
        df_line_details = df_line_details.groupby(['po_os_ref', 'po_position', 'po_level', 'rate_diff_percent']). \
            size().reset_index()
        df_line_details.columns = ['po_os_ref', 'po_position', 'po_level', 'rate_diff_percent', 'count']
        df_line_details = df_line_details.pivot_table(index=['po_position', 'po_os_ref', 'rate_diff_percent'],
                                                      columns='po_level',
                                                      values='count')
        rs_summary = df_line_details.reset_index()
        rs_summary.sort_values(by=['po_os_ref', 'po_position', 'rate_diff_percent'],
                               ascending=[True, True, True],
                               inplace=True)
        rs_summary = rs_summary.fillna(0.0).to_dict('records')

        df_unit_price = pd.DataFrame(list(self.qs_unit_price.values()))
        df_unit_price = df_unit_price.pivot(index='po_position', columns='po_level', values='amount').reset_index()
        rs_unit_price = df_unit_price.to_dict('records')

        context = {
            'contractor': self.qs_po.contractor,
            'po_num': self.qs_po.po_num,
            'po_line_count': self.qs_po.line,
            'rs_resource': rs_resource,  # recordset for individual resource (employee information)
            'rs_summary': rs_summary,  # recordset for summary of total numbers
            'rs_unit_price': rs_unit_price,  # recordset for summary of unit price
            'period_string': self.period_string,
            'total_required_hour': self.required_hour
        }

        return context

    def single_set_sbh(self, context=None):

        wb = load_workbook(self.form)

        # WORK SHEET SBH-FORM 1
        # ---------------------------------------------------------------------------------------
        row = 15
        column = 2
        sr = 1
        ws = wb.get_sheet_by_name('SBH-FORM 1')

        ws.cell('D5').value = context['contractor']
        ws.cell('G5').value = context['po_num']
        ws.cell('G7').value = 1
        ws.cell('D7').value = context['period_string']

        rs_resource = context['rs_resource']

        ws.insert_rows(row, len(rs_resource) - 1)

        for record in rs_resource:
            ws.cell(row=row, column=column).value = sr
            ws.cell(row=row, column=column + 1).value = record['po_os_ref']
            ws.cell(row=row, column=column + 2).value = record['agency_ref_num']
            ws.cell(row=row, column=column + 3).value = record['res_full_name']
            ws.cell(row=row, column=column + 4).value = record['po_position']
            ws.cell(row=row, column=column + 5).value = record['po_level']
            ws.cell(row=row, column=column + 6).value = '{0:%d-%b-%Y}'.format(record['date_of_join'])
            ws.cell(row=row, column=column + 7).value = record['po_line_detail__rate_diff_percent']
            ws.cell(row=row, column=column + 8).value = record['required_hour']
            ws.cell(row=row, column=column + 9).value = record['invoice_claim']
            ws.cell(row=row, column=column + 10).value = ''
            ws.cell(row=row, column=column + 11).value = record['remarks']
            ws.cell(row=row, column=column + 12).value = record['po_line_detail__director_name']
            ws.cell(row=row, column=column + 13).value = ''

            # # Unlock Cells
            ws.cell(row=row, column=column + 9).protection = Protection(locked=False)
            ws.cell(row=row, column=column + 11).protection = Protection(locked=False)
            ws.cell(row=row, column=column + 12).protection = Protection(locked=False)
            ws.cell(row=row, column=column + 13).protection = Protection(locked=False)

            sr += 1
            row += 1
        # ---------------------------------------------------------------------------------------
        # END SBH FORM 1


        # WORK SHEET SBH-FORM 2
        # ---------------------------------------------------------------------------------------
        row = 16
        column = 2  # B

        ws = wb.get_sheet_by_name('SBH-FORM 2')

        ws.insert_rows(row, len(context['rs_summary']) - 1)
        ws.cell('R13').value = context['total_required_hour']

        for record in context['rs_summary']:
            ws.cell(row=row, column=column).value = record['po_os_ref']
            ws.cell(row=row, column=column + 1).value = record.get('po_position')
            ws.cell(row=row, column=column + 2).value = record.get('Level 1', 0)
            ws.cell(row=row, column=column + 3).value = record.get('Level 2', 0)
            ws.cell(row=row, column=column + 4).value = record.get('Level 3', 0)
            ws.cell(row=row, column=column + 12).value = record.get('rate_diff_percent', 0)

            row += 1
        # ---------------------------------------------------------------------------------------
        # END SBH FORM 2

        # WORK SHEET UNIT PRICE
        # ---------------------------------------------------------------------------------------
        row = 1
        column = 1  # B

        ws = wb.get_sheet_by_name('UNIT PRICE')

        for record in context['rs_unit_price']:
            ws.cell(row=row, column=column).value = record['po_position']
            ws.cell(row=row, column=column + 1).value = record.get('Level 1', 0)
            ws.cell(row=row, column=column + 2).value = record.get('Level 2', 0)
            ws.cell(row=row, column=column + 3).value = record.get('Level 3', 0)
            ws.cell(row=row, column=column + 4).value = record.get('Level 4', 0)

            row += 1
        # ---------------------------------------------------------------------------------------
        # END UNIT PRICE

        for sheet in wb.worksheets:
            sheet.protection.enable()
            sheet.protection.set_password('tbpc19')
        wb.save(os.path.join(self.base_dir, 'output', self.output_file))

    def make_sbh_per_po(self, po_num=None, period_start=None, period_end=None,
                        ramadan_start=None, ramadan_end=None):
        self.po_num = po_num
        self.period_start = period_start
        self.period_end = period_end
        self.ramadan_start = ramadan_start
        self.ramadan_end = ramadan_end
        self.set_variables()
        self.set_initial_queryset()
        try:
            context = self.get_context()
            self.output_file = get_output_file(
                '{:0>2}'.format(self.period_start.month),
                context['contractor'],
                context['po_num'],
                context['period_string'],
            )
            self.single_set_sbh(context)
        except Exception as e:
            print('Failed: ', po_num, ' due to missing {}'.format(str(e)))

    def make_sbh_per_contractor(self, contractor=None, period_start=None, period_end=None,
                                ramadan_start=None, ramadan_end=None):
        if contractor.lower() == 'all':
            qs = PurchaseOrder.objects.all()
        else:
            qs = PurchaseOrder.objects.filter(contractor=contractor).all()

        for i in qs:
            self.make_sbh_per_po(i.po_num, period_start, period_end, ramadan_start, ramadan_end)

    def make_sbh_per_division(self, po_num=None, division=None, period_start=None, period_end=None,
                              ramadan_start=None, ramadan_end=None):
        self.division = division
        self.po_num = po_num
        self.period_start = period_start
        self.period_end = period_end
        self.ramadan_start = ramadan_start
        self.ramadan_end = ramadan_end
        self.set_variables()
        self.set_initial_queryset()
        # try:
        context = self.get_context()
        self.output_file = get_output_file(
            '{:0>3}'.format(self.period_start.month),
            self.division or '',
            context['contractor'],
            context['po_num'],
            context['period_string'],
        )
        self.single_set_sbh(context)
        # except Exception as e:
        #     print('Failed: ', po_num, ' due to missing {}'.format(str(e)))


class DATASHEET(SBH):
    def __init__(self):
        self.contractor = None
        super(DATASHEET, self).__init__()
        self.form = os.path.join(self.base_dir, 'forms', 'DATASHEET FOR CONTRACTOR.xlsx')

    def set_initial_queryset(self):
        q = Q(contractor=self.contractor)
        self.qs_resource = Resource.objects.filter(q)

    def get_context(self):
        df_resource = pd.DataFrame(list(self.qs_resource.values()))
        df_resource.sort_values(
            by=['division', 'manager', 'res_job_title'],
            ascending=[True, True, True],
            inplace=True)
        rs_resource = df_resource.to_dict('records')

        context = {
            'contractor': self.contractor,
            'rs_resource': rs_resource,  # recordset for individual resource (employee information)
            'period_string': self.period_string,
        }

        return context

    def single_set_datasheet(self, context=None):

        wb = load_workbook(self.form)

        # WORK SHEET SBH-FORM 1
        # ---------------------------------------------------------------------------------------
        row = 7
        column = 2
        sr = 1
        ws = wb.get_sheet_by_name('STAFF LIST')

        ws.cell('B1').value = "713H CLAIMS FOR {}".format(context['contractor'])
        ws.cell('B2').value = context['period_string']

        rs_resource = context['rs_resource']

        ws.insert_rows(row, len(rs_resource) - 1)

        for record in rs_resource:
            ws.cell(row=row, column=column).value = sr
            ws.cell(row=row, column=column + 1).value = record.get('division', '')
            ws.cell(row=row, column=column + 2).value = record.get('section', '')
            ws.cell(row=row, column=column + 3).value = record.get('manager', '')
            ws.cell(row=row, column=column + 4).value = record.get('id', '')
            ws.cell(row=row, column=column + 5).value = record.get('po_line_detail_id', '')
            ws.cell(row=row, column=column + 6).value = record.get('rate', '')
            ws.cell(row=row, column=column + 7).value = record.get('agency_ref_num', '')
            ws.cell(row=row, column=column + 8).value = record.get('res_full_name', '').title()
            ws.cell(row=row, column=column + 9).value = record.get('res_job_title', '')
            ws.cell(row=row, column=column + 10).value = record.get('grade_level', '')
            ws.cell(row=row, column=column + 11).value = '{0:%d-%b-%Y}'.format(record.get('date_of_join', ''))
            ws.cell(row=row, column=column + 12).value = 'Yes' if record.get('has_tool_or_uniform', '') else ""
            ws.cell(row=row, column=column + 13).value = self.required_hour

            # # Unlock Cells
            ws.cell(row=row, column=column + 14).protection = Protection(locked=False)
            ws.cell(row=row, column=column + 15).protection = Protection(locked=False)
            ws.cell(row=row, column=column + 16).protection = Protection(locked=False)

            sr += 1
            row += 1
        # ---------------------------------------------------------------------------------------
        # END SBH FORM 1

        for sheet in wb.worksheets:
            sheet.protection.enable()
            sheet.protection.set_password('tbpc19')
        wb.save(os.path.join(self.base_dir, 'output', self.output_file))

    def make_datasheet(self, contractor=None, period_start=None, period_end=None,
                       ramadan_start=None, ramadan_end=None):
        self.contractor = contractor
        self.period_start = period_start
        self.period_end = period_end
        self.ramadan_start = ramadan_start
        self.ramadan_end = ramadan_end
        self.set_variables()
        self.set_initial_queryset()
        try:
            context = self.get_context()
            self.output_file = get_output_file(
                'DATASHEET',
#                '{:0>2}'.format(self.period_start.month),
                context['contractor'],
                context['period_string']
            )
            self.single_set_datasheet(context)
        except Exception as e:
            print('Failed: ', contractor, ' due to missing {}'.format(str(e)))


    def make_datasheet_per_contractor(self, contractor=None, period_start=None, period_end=None,
                                      ramadan_start=None, ramadan_end=None):
        if contractor.lower() == 'all':
            contractors = list(Resource.objects.values_list('contractor', flat=True).distinct())
            for contractor in contractors:
                self.make_datasheet(contractor, period_start, period_end,
                                    ramadan_start, ramadan_end)
        else:
            self.make_datasheet(contractor, period_start, period_end,
                                ramadan_start, ramadan_end)


if __name__ == '__main__':
    import timeit, functools
    sbh = SBH()
    datasheet = DATASHEET()
#    t = timeit.Timer(functools.partial(datasheet.make_datasheet_per_contractor, 'all', '20/07/2016', '19/08/2016'))
    t = timeit.Timer(functools.partial(sbh.make_sbh_per_contractor, 'all', '20/07/2016', '19/08/2016', None, None))
    print(t.timeit(1))
    # datasheet = DATASHEET()
    #datasheet.make_datasheet_per_contractor('all', '20/07/2016', '19/08/2016')

    # sbh = SBH()
    # sbh.make_sbh_per_division('1072656-0', 'CSE', '20/05/2016', '19/06/2016', '5/6/2016', '19/06/2016')
    # sbh.make_sbh_per_division('1072658-0', 'CSE', '20/05/2016', '19/06/2016', '5/6/2016', '19/06/2016')
    # sbh.make_sbh_per_division('1072659-0', 'CSE', '20/05/2016', '19/06/2016', '5/6/2016', '19/06/2016')
    # sbh.make_sbh_per_po('1070509-0', '20/05/2016', '19/06/2016', '5/6/2016', '19/06/2016')
    # sbh.make_sbh_per_po('1068851-0', '20/05/2016', '19/06/2016', '5/6/2016', '19/06/2016')
    # # print(get_required_hour(to_date_format('20/05/2016'), to_date_format('19/06/2016'), to_date_format('5/6/2016'), to_date_format('19/06/2016')))
    # # print(get_required_hour(to_date_format('20/05/2016'), to_date_format('19/06/2016')))
